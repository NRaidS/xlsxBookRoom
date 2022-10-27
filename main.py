import openpyxl
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
import matplotlib.pyplot as plt
from pandas.plotting import table
# 解决 画图中文 方块问题
from pylab import mpl

res = []    #缓存个人已约信息

def init():
    print("欢迎光~临！")

#初始化本人预约信息
def inityue(instu):
    wb = load_workbook('clstme.xlsx')
    # 打开sheet, (一般打开一个已经存在的文件后，我们希望直接操作其中的某个表格)
    # 我们可以get所有表格的名字，再从中选择某个表格
    sheet_list = wb.sheetnames
    ws = wb[sheet_list[0]]
    # 或者直接用表格的名字
    ws = wb['Sheet1']
    for r in range(2, ws.max_row+1):
        for c in range(2, ws.max_column+1): #深入到每一个单元格子
            #ws.cell(r, c).value.splitlines()
            vle = str(ws.cell(r, c).value).splitlines() #一个单元格子有若干行，将每行作为一个列表对象
            for i in range(len(vle)):#到此，即是分析每一个预约数据！vle[i]
                if instu in vle[i]:
                    res.append([vle[i],"星期：" + str(c - 1),"时段：" + str(r - 1)])

#登录界面
def comin():
    print("请输入您的学号：（输入0退出，测试账号为114514,姓名：IKUN）")
    inno = input()  #获得学号
    if inno == "0":
        return 1    #退出返回1
    else:
        print("请输入您的姓名：")
        instu = input() #获得姓名

        #获取工作簿对象
        wb=openpyxl.load_workbook("stuNo.xlsx")
        #获得工作簿的表名后，就可以获得表对象
        ws=wb["Sheet1"]
        #获取该表相应的行数
        ros=ws.max_row

        #for判断是否在学生库中
        for c in range(1,ros + 1):
            if instu == ws.cell(row = c,column = 2).value and inno == ws.cell(row=c,column=1).value:
                inityue(instu)#缓存本人预约信息
                #query(instu)
                menu(inno,instu)#跳转到菜单界面
                return 1    #进入菜单后退出，直接退出系统
        print("该用户不存在！请检查输入是否正确！")
        return 0    #重新输入

#菜单界面
def menu(inno, instu):
    while 1:    #重复询问下次操作
        print("请输入操作：")
        print("1：预约实验室")
        print("2：显示目前全部的预约信息")
        print("3：查询所有教室")
        print("4：查询预约情况")
        print("5：修改预约信息")
        print("6：退出系统")
        x = input()
        if x == '1':
            while booking(instu):   #当booking返回1时，表明没运行到底，继续运行
                continue
        elif x == '2': show() #显示目前全部的预约信息
        elif x == '3': sowrom()   #显示所有教室
        elif x == '4': query()    #查询本人已约信息
        elif x == '5': rebook(instu)  #修改本人预约信息
        elif x == '6':
            exitt() #退出
            break   #break循环访问操作，退出系统
        else:
            print("输入错误！请重新输入\n")

#预约教室
def booking(instu):
    print("请输入需预约的教室：(退出请按0，输出所有教室请按1)")
    room = input()
    if(room == '0'):
        return 0    #停止运行
    if(room == '1'):
        sowrom()
        return 1    #继续运行

    #下面判断教室是否存在
    wb=openpyxl.load_workbook("classInfo.xlsx")
    #获得工作簿的表名后，就可以获得表对象
    ws=wb["Sheet"]
    #获取该表相应的行数
    ros=ws.max_row
    flag = 0
    for i in range(1,ros + 1):
        if ws.cell(row = i,column = 1).value == room:
            flag = 1
            break   #找到直接break，无需继续找
    if flag == 0:
        print("教室不存在！请重新输入")
        return 1

    print("请选择星期：（直接输入对应数字）")
    day = input()
    #input默认输入格式为str，非整数直接转str会报错。需先判断是否能转。
    if not day.isnumeric():
        print("星期输入有误！请重新输入!")
        return 1
    day = int(day)
    if day not in range(1,8):
        print("星期输入有误！请重新输入")
        return 1
    else:
        print("请输入时段：（1上午、2下午、3晚上）")
        tme = input()
        #input默认输入格式为str，非整数直接转int会报错。需先判断是否能转。
        if not tme.isnumeric():
            print("时段输入有误！请重新输入!")
            return 1
        tme = int(tme)
        if tme not in range(1,4):
            print("时段输入有误！请重新输入")
            return 1
        #开始预约
        else:
            wb = load_workbook('clstme.xlsx')
            # 打开sheet, (一般打开一个已经存在的文件后，我们希望直接操作其中的某个表格)
            # 我们可以get所有表格的名字，再从中选择某个表格
            sheet_list = wb.sheetnames
            ws = wb[sheet_list[0]]
            # 或者直接用表格的名字
            ws = wb['Sheet1']
            vle = str(ws.cell(tme + 1, day + 1).value).splitlines() #一个单元格子有若干行，将每行作为一个列表对象
            for i in range(len(vle)):#到此，即是分析每一个预约数据！vle[i]
                if room in vle[i]:
                    print("该时段已有人预约！请重新输入！")
                    return 1

            #在clstme表中加上预约信息
            cel = str(ws.cell(tme + 1, day + 1).value)  #获得原来的单元格值
            nul = str(ws.cell(1, 1).value)  #获得空白单元格做对比
            #print(cel,len(cel))
            if cel == nul:
                cel = str(room) + "," + str(instu)  #单元格为空，没必要加换行符
            else:
                cel += "\n" + str(room) + "," + str(instu)  #单元格已有数据，在后面加上换行符
            ws.cell(tme + 1,day + 1,cel)    #赋新值
            wb.save(filename='clstme.xlsx')

            #更新预约缓存
            res.append([str(room +  "," + instu), "星期：" + str(day), "时段：" + str(tme)])

            print("您已预约成功！")
            return 0

#显示全部教室
def sowrom():
    wb=openpyxl.load_workbook("classInfo.xlsx")
    #获得工作簿的表名后，就可以获得表对象
    ws=wb["Sheet"]
    #获取该表相应的行数
    ros=ws.max_row
    for i in range(1,ros + 1):
        print(ws.cell(row = i,column = 1).value)

#显示全部预约信息（以图片的方式）
def show():
    # #直接打开文件展示，有着安全风险。
    # file = "clstme.xlsx"
    # os.startfile(file)

    mpl.rcParams['font.sans-serif'] = ['Microsoft YaHei']    # 指定默认字体：解决plot不能显示中文问题
    mpl.rcParams['axes.unicode_minus'] = False
    # figsize 指定figure的宽和高，单位为英寸；
    # dpi参数指定绘图对象的分辨率，即每英寸多少个像素，缺省值为80      1英寸等于2.5cm,A4纸是 21*30cm的纸张
    fig = plt.figure(figsize=(9, 10), dpi=900)
    # frameon:是否显示边框
    ax = fig.add_subplot(161, frame_on=False,)
    # 隐藏x轴 y轴
    ax.xaxis.set_visible(False)  # hide the x axis
    ax.yaxis.set_visible(False)  # hide the y axis

    datas = pd.read_excel('clstme.xlsx')
    datas = datas.iloc[:, 0:]

    print(datas)
    print("请等待图片生成...")
    # 生成图片
    table(ax, datas, loc='center')  # where df is your data frame
    # 保存图片
    plt.savefig('photo.jpg')
    os.startfile("photo.jpg")
    return

#查询本人已约信息
def query():
    for i in range(len(res)):
        print(str(i + 1) + "、" + str(res[i]))
    print("时段：（1上午、2下午、3晚上）")
    return

#重约教室
def rebook(instu):
    if len(res) == 0:
        print("您当前没有预约教室，请先预约！")
        return
    else:
        print("请输入你想修改哪一条预约信息：")
        print("[""str"",""x"",""y""]表示在星期x时段y时已预约str教室")
        query()

        x = int(input())
        if x not in range(1, len(res) + 1):
            print("输入有误！，请重新输入")
            rebook(instu)
        else:
            x = x - 1
            days = int(res[x][1][len(res[x][1]) - 1])
            tmes = int(res[x][2][len(res[x][2]) - 1])
            roms = ""
            for i in range(len(res[x][0])):###
                if res[x][0][i] != ',':
                    roms += res[x][0][i]
                else:
                    break

            #print(days,tmes,roms)
            while booking(instu):
                continue

            #把原来的信息删除
            wb = load_workbook('clstme.xlsx')
            # 打开sheet, (一般打开一个已经存在的文件后，我们希望直接操作其中的某个表格)
            # 我们可以get所有表格的名字，再从中选择某个表格
            sheet_list = wb.sheetnames
            ws = wb[sheet_list[0]]
            # 或者直接用表格的名字
            ws = wb['Sheet1']
            yuan = ws.cell(tmes + 1, days + 1).value
            if "\n" in yuan:
                shan = "\n" + roms + "," + instu
            else:
                shan = roms + "," + instu
            #print(yuan,type(yuan),shan,type(shan))
            xin = yuan.replace(shan, '')
            if xin == yuan:
                shan = roms + "," + instu + "\n"
            xin = yuan.replace(shan, '')
            ws.cell(tmes + 1, days + 1,xin)
            wb.save(filename='clstme.xlsx')

            #更改预约缓存
            tp = roms + "," + instu
            for i in range(len(res)):
                if tp in res[i]:
                    res.pop(i)
                    break

            return
#退出函数
def exitt():
    print("栓Q")

#主程序
while comin() == 0:
    continue

